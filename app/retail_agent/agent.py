"""Retail Agent — async LangChain wrapper.

Reads the buylead's MCAT, title, quantity, and ISQs, and asks the LLM to
classify whether the buyer is a retailer or end-consumer.

LLM output contract (from prompt.md):
    {
        "Display_id": str,
        "Classification": "RETAIL" | "NON-RETAIL" | "UNCLASSIFIED",
        "Classi_Score": float (0.0–1.0),
        "Confidence": "High" | "Medium" | "Low" | "None",
        "Override_Applied": bool,
        "Reason": str
    }
"""
import asyncio
import csv
import json
import logging
import os
import re
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, TypedDict

from langchain_core.messages import HumanMessage, SystemMessage
from langchain_openai import ChatOpenAI
from langgraph.graph import END, StateGraph
from openpyxl import load_workbook

log = logging.getLogger("bl-auditor.retail_agent")
_LLM_MAX_RETRIES = int(os.getenv("LLM_MAX_RETRIES", "2"))


ROOT_DIR = Path(__file__).resolve().parents[2]
PROMPT_PATH = Path(__file__).resolve().with_name("prompt.md")
MCAT_DATA_PATH = ROOT_DIR / "mcat_data.xlsx"
EVIDENCE_DATA_PATH = ROOT_DIR / "evidence_data2.csv"


class RetailState(TypedDict, total=False):
    offer_id: str
    buylead_response: Dict[str, Any]
    agent_input: Dict[str, Any]
    system_prompt: str
    user_message: str
    raw_output: str
    result: Dict[str, Any]


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    return "" if text in {"null", "undefined"} else text


def _to_number(value: Any, default: float = 0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _parse_jsonish(raw: Any) -> Any:
    if raw in (None, ""):
        return None
    if not isinstance(raw, str):
        raw = json.dumps(raw)
    raw = raw.replace('\\"', '"')
    parsed = json.loads(raw)
    if isinstance(parsed, str):
        parsed = json.loads(parsed)
    return parsed


_UNIT_SYNONYMS: Dict[str, str] = {
    # Tonne family — mcat uses TONNE, not TON
    "TON": "TONNE", "TONS": "TONNE", "TONE": "TONNE", "TONES": "TONNE",
    "TONNES": "TONNE", "TONNE": "TONNE",
    # Litre family
    "LITER": "LITRE", "LITERS": "LITRE", "LITRES": "LITRE", "LITRE": "LITRE",
    "LTR": "LITRE", "LTRS": "LITRE",
    # Piece family
    "PIECES": "PIECE", "PIECE": "PIECE", "PCS": "PIECE", "PC": "PIECE",
    # Box family — fixes "Boxes" → "BOX"
    "BOXES": "BOX", "BOX": "BOX",
    # Bottle family
    "BOTTLES": "BOTTLE", "BOTTLE": "BOTTLE",
    # Kilogram family
    "KILO": "KG", "KILOS": "KG", "KILOGRAM": "KG", "KILOGRAMS": "KG",
    "KGS": "KG", "KG": "KG",
    # Gram family — mcat uses GRAM, not G
    "GRAM": "GRAM", "GRAMS": "GRAM", "GMS": "GRAM", "GM": "GRAM", "G": "GRAM",
    # Bag, Jar, Set, Pack, Carton, Case, Pair, Roll
    "BAGS": "BAG", "BAG": "BAG",
    "JARS": "JAR", "JAR": "JAR",
    "SETS": "SET", "SET": "SET",
    "PACKS": "PACK", "PACK": "PACK",
    "CARTONS": "CARTON", "CARTON": "CARTON",
    "CASES": "CASE", "CASE": "CASE",
    "PAIRS": "PAIR", "PAIR": "PAIR",
    "ROLLS": "ROLL", "ROLL": "ROLL",
    # Meter family
    "METER": "METER", "METERS": "METER", "METRE": "METER", "METRES": "METER",
}


def _canonical_unit(value: Any) -> str:
    """Map a unit string (BL qty or evidence-file column) to a canonical
    uppercase form aligned with mcat's ``unit_display_name``.

    Strips leading numeric prefix (e.g. "50 Piece" → "Piece") and applies a
    synonym table for known typos/plurals. Multi-word phrases like "Carton
    Of 100 Pieces" pass through uppercased verbatim so they can match mcat
    entries such as "BOX OF 100 PIECES" if present.
    """
    raw = str(value or "").strip()
    raw = re.sub(r"^\d+\s*", "", raw).strip()
    if not raw:
        return ""
    upper = raw.upper()
    if " " in upper:
        return upper
    return _UNIT_SYNONYMS.get(upper, upper)


def _normalize_unit_from_qty(qty: str) -> str:
    return _canonical_unit(qty)


def _normalize_mcat_unit(value: Any) -> str:
    return "" if value is None else str(value).strip().upper()


def _extract_qty(qty_str: Any) -> float | None:
    if not qty_str:
        return None
    match = re.search(r"\d+", str(qty_str))
    return float(match.group(0)) if match else None


def _get_slab(qty: float | None) -> str:
    if qty is None or qty <= 0:
        return "no_slab"
    if qty <= 10:
        return "1-10"
    if qty <= 20:
        return "11-20"
    if qty <= 30:
        return "21-30"
    if qty <= 50:
        return "31-50"
    if qty <= 75:
        return "51-75"
    if qty <= 100:
        return "76-100"
    if qty <= 150:
        return "101-150"
    if qty <= 200:
        return "151-200"
    if qty <= 300:
        return "201-300"
    return ">300"


_BULK_UNITS = {"TON", "TONNE", "MT", "METRIC TON", "QUINTAL", "KL", "KILOLITRE"}


def _hard_override(offer: Dict[str, Any], qty: float | None) -> Dict[str, Any] | None:
    unit = _canonical_unit(offer.get("Qty"))
    if unit in _BULK_UNITS:
        rule = f"Bulk unit ({unit})"
    elif unit == "KG" and qty is not None and qty >= 200:
        rule = "KG ≥ 200"
    elif unit == "LITRE" and qty is not None and qty >= 200:
        rule = "Litre ≥ 200"
    else:
        return None
    return {
        "Display_id": offer.get("Display_id"),
        "Classification": "NON-RETAIL",
        "Classi_Score": 1.0,
        "Confidence": "High",
        "Override_Applied": f"Yes — {rule}",
        "Reason": f"Hard override: {rule} is always non-retail.",
    }


def _read_prompt() -> str:
    from app.services.prompt_override_service import get_active_prompt
    return get_active_prompt("retail")[0]


def _render_template(template: str, data: Dict[str, Any]) -> str:
    def replace(match: re.Match[str]) -> str:
        key = match.group(1).strip()
        value = data.get(key, "")
        if isinstance(value, str):
            return value
        return json.dumps(value, ensure_ascii=False)

    return re.sub(r"{{\s*([^}]+)\s*}}", replace, template)


def _build_offer_source(offer_id: str, buylead_response: Dict[str, Any]) -> Dict[str, Any]:
    data = buylead_response.get("RESPONSE", {}).get("DATA", {})
    isq_info = data.get("ENRICHMENTINFO")

    buyer_viewed = [
        {
            "eto_ofr_buyer_sell_mcats": data.get("ETO_OFR_BUYER_SELL_MCATS"),
            "eto_ofr_buyer_prime_mcats": data.get("ETO_OFR_BUYER_PRIME_MCATS"),
        }
    ]

    bl_card_data: List[Dict[str, Any]] = []
    for product in data.get("PRODUCTS_ENQUIRED") or []:
        if isinstance(product, dict):
            bl_card_data.append(product)

    return {
        "Display_id": data.get("ETO_OFR_DISPLAY_ID") or offer_id,
        "Title": data.get("ETO_OFR_TITLE") or "",
        "MCAT": data.get("PRIME_MCAT_NAME") or data.get("ETO_OFR_GLCAT_MCAT_NAME") or "",
        "MCAT_id": data.get("FK_GLCAT_MCAT_ID") or "",
        "ISQ_info": isq_info,
        "BL_Type": data.get("BY_LEAD_TYPE") or data.get("FK_ETO_OFR_TYPE_ID"),
        "Buyer_viewed": buyer_viewed,
        "BL_card_data": bl_card_data,
    }


def _bl_detail_and_keys(source: Dict[str, Any]) -> Dict[str, Any]:
    isq_parsed: Dict[str, str] = {}
    qty = ""
    order_value = ""

    try:
        parsed = _parse_jsonish(source.get("ISQ_info"))
        if isinstance(parsed, dict):
            parsed = parsed.get("1", [])
        if isinstance(parsed, list):
            for question in parsed:
                key = str(question.get("DESC") or "").strip() if isinstance(question, dict) else ""
                value = str(question.get("RESPONSE") or "").strip() if isinstance(question, dict) else ""
                if not key:
                    continue
                lowered = key.lower()
                if lowered == "quantity":
                    qty = value
                    continue
                if "quantity" in lowered and not qty:
                    qty = value
                    continue
                if "order value" in lowered:
                    order_value = value
                    continue
                isq_parsed[key] = value
    except Exception as exc:
        log.warning("retail_agent ISQ parse failed (qty/unit will be empty): %s", exc)

    unit = _normalize_unit_from_qty(qty)
    mcat_id = source.get("MCAT_id") or ""
    mcat_unit = f"{mcat_id}-{unit}" if mcat_id and unit else ""

    retail_flag = "No"
    try:
        if int(source.get("BL_Type") or 0) in {1, 3, 5, 6}:
            retail_flag = "Yes"
    except (TypeError, ValueError):
        pass

    buyer_obj: Dict[str, Any] = {}
    try:
        parsed_buyer = _parse_jsonish(source.get("Buyer_viewed"))
        if isinstance(parsed_buyer, list):
            buyer_obj = parsed_buyer[0] or {}
        elif isinstance(parsed_buyer, dict):
            values = list(parsed_buyer.values())
            buyer_obj = values[0] if values and isinstance(values[0], dict) else parsed_buyer
    except Exception:
        pass

    bl_card = []
    try:
        parsed_card = _parse_jsonish(source.get("BL_card_data"))
        if isinstance(parsed_card, dict):
            parsed_card = [parsed_card]
        if isinstance(parsed_card, list):
            for card in parsed_card:
                if not isinstance(card, dict):
                    continue
                item_name = str(card.get("FK_PC_ITEM_NAME") or card.get("FK_PC_ITEM_DISPLAY_NAME") or "").strip()
                if not item_name:
                    continue
                price = str(card.get("PRODUCT_PRICE") or "").replace("\\u20b9", "₹").strip()
                bl_card.append({"Item Name": item_name, "Price": price})
    except Exception:
        bl_card = []

    return {
        "Display_id": source.get("Display_id"),
        "Title": source.get("Title"),
        "MCAT": source.get("MCAT"),
        "MCAT_id": mcat_id,
        "MCAT_Unit": mcat_unit,
        "Retail_Flag": retail_flag,
        "ISQ": isq_parsed,
        "Qty": qty,
        "Order_Value": order_value,
        "Buyer_viewed": {
            "Sells": _clean(buyer_obj.get("eto_ofr_buyer_sell_mcats")) or "No Selling Activity",
            "Buys": _clean(buyer_obj.get("eto_ofr_buyer_prime_mcats")),
        },
        "BL_card": bl_card,
    }


_CSV_METRIC_MAP = {
    "bl_apprvd": "bl_apprvd",
    "pur": "pur",
    "pur_retailer": "pur_retailer",
    "pur_wholesaler": "pur_wholesaler",
    "retail_ni": "retail_ni",
    "ni_retailer": "ret_ni_cnt_retailer",
    "ni_wholesaler": "ret_ni_cnt_wholesaler",
}


def _empty_evidence_metrics() -> Dict[str, Any]:
    return {
        "bl_apprvd": 0, "pur": 0, "pur_retailer": 0, "pur_wholesaler": 0,
        "retail_ni": 0, "ni_retailer": 0, "ni_wholesaler": 0,
        "bucket_count": 0,
    }


@lru_cache(maxsize=1)
def _load_evidence_metrics() -> tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    """Build two evidence indexes from evidence_data2.csv (pre-aggregated):
       - by_slab:  key = f"{norm_unit}_{slab}" → metrics for that exact bucket
       - by_unit:  key = f"{norm_unit}"        → metrics summed across all slabs

    The CSV's MCAT_Unit column is rebuilt from glcat_mcat_id + canonicalized
    eto_ofr_qty_unit so lookup keys match the runtime side exactly (which also
    runs unit through _canonical_unit).
    """
    if not EVIDENCE_DATA_PATH.exists():
        return {}, {}

    by_slab: Dict[str, Dict[str, Any]] = {}
    by_unit: Dict[str, Dict[str, Any]] = {}

    with EVIDENCE_DATA_PATH.open("r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            mcat_id_raw = (row.get("glcat_mcat_id") or "").strip()
            if not mcat_id_raw:
                continue
            try:
                mcat_id_text = str(int(float(mcat_id_raw)))
            except ValueError:
                mcat_id_text = mcat_id_raw

            raw_unit = (row.get("eto_ofr_qty_unit") or "").strip()
            canonical_unit = _canonical_unit(raw_unit)
            display_unit = canonical_unit if canonical_unit else "No_Unit"
            norm_unit = _normalize_mcat_unit(f"{mcat_id_text}-{display_unit}")
            if not norm_unit:
                continue

            slab = (row.get("qty_slab") or "").strip() or "no_slab"
            slab_key = f"{norm_unit}_{slab}"
            unit_key = norm_unit

            mcat_name = (row.get("glcat_mcat_name") or "").strip() or "Unknown"

            if slab_key not in by_slab:
                entry = _empty_evidence_metrics()
                entry["glcat_mcat_id"] = mcat_id_text
                entry["glcat_mcat_name"] = mcat_name
                entry["eto_ofr_qty_unit"] = raw_unit
                entry["slab"] = slab
                entry["MCAT_Unit"] = norm_unit
                by_slab[slab_key] = entry

            if unit_key not in by_unit:
                entry = _empty_evidence_metrics()
                entry["glcat_mcat_id"] = mcat_id_text
                entry["glcat_mcat_name"] = mcat_name
                entry["eto_ofr_qty_unit"] = raw_unit
                entry["MCAT_Unit"] = norm_unit
                by_unit[unit_key] = entry

            for out_field, src_col in _CSV_METRIC_MAP.items():
                val = _to_number(row.get(src_col))
                by_slab[slab_key][out_field] += val
                by_unit[unit_key][out_field] += val

            by_slab[slab_key]["bucket_count"] += 1
            by_unit[unit_key]["bucket_count"] += 1

    return by_slab, by_unit


@lru_cache(maxsize=1)
def _load_price_index() -> Dict[str, Dict[str, float]]:
    """Build map from canonical `{mcat_id}-{UNIT}` → q1/median/q3 once.

    mcat_data.xlsx has duplicate 'median'/'q3' header names. This uses the
    FIRST occurrence (positions 8, 9, 10 in current schema) to stay
    deterministic if the duplicate columns ever diverge.

    Returns an empty index if mcat_data.xlsx is absent so the agent can still
    classify without price quartiles.
    """
    if not MCAT_DATA_PATH.exists():
        return {}
    wb = load_workbook(MCAT_DATA_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))

    def _first(name: str) -> int:
        for i, h in enumerate(headers):
            if h == name:
                return i
        return -1

    mcat_idx = _first("fk_glcat_mcat_id")
    unit_idx = _first("unit_display_name")
    q1_idx = _first("q1")
    median_idx = _first("median")
    q3_idx = _first("q3")

    index: Dict[str, Dict[str, float]] = {}
    if -1 in (mcat_idx, unit_idx, q1_idx, median_idx, q3_idx):
        wb.close()
        return index

    for row in rows:
        mcat_id = row[mcat_idx]
        unit = row[unit_idx]
        if mcat_id in (None, "") or unit in (None, ""):
            continue
        mcat_id_text = str(int(mcat_id)) if isinstance(mcat_id, float) else str(mcat_id)
        key = _normalize_mcat_unit(f"{mcat_id_text}-{unit}")
        if key in index:
            continue
        index[key] = {
            "q1": _to_number(row[q1_idx]),
            "median": _to_number(row[median_idx]),
            "q3": _to_number(row[q3_idx]),
        }
    wb.close()
    return index


def _find_price_data(norm_unit: str) -> Dict[str, float] | None:
    if not norm_unit:
        return None
    return _load_price_index().get(norm_unit)


def _clean_classifier_output(raw: str) -> Dict[str, Any]:
    text = (raw or "").replace("```json", "").replace("```", "").strip()
    if not text:
        raise ValueError("LLM returned empty response")
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{[\s\S]*\}", text)
        if not match:
            raise ValueError(f"No JSON object found in LLM output: {text[:300]}")
        return json.loads(match.group(0))


async def _prepare_input(state: RetailState) -> RetailState:
    offer_id = state["offer_id"]
    source = _build_offer_source(offer_id, state["buylead_response"])
    offer = _bl_detail_and_keys(source)

    qty = _extract_qty(offer.get("Qty"))
    slab = _get_slab(qty)
    norm_unit = _normalize_mcat_unit(offer.get("MCAT_Unit"))

    by_slab, by_unit = _load_evidence_metrics()
    metrics: Dict[str, Any] | None = None
    evidence_match = "no_data"
    slab_key = f"{norm_unit}_{slab}" if (norm_unit and slab and slab != "no_slab") else ""
    if slab_key:
        metrics = by_slab.get(slab_key)
        if metrics is not None:
            evidence_match = "exact"
    if metrics is None and norm_unit:
        metrics = by_unit.get(norm_unit)
        if metrics is not None:
            evidence_match = "unit_only"
    if metrics is None:
        metrics = _empty_evidence_metrics()

    price = _find_price_data(norm_unit)
    if price is None:
        price = {"q1": 0, "median": 0, "q3": 0}
        price_match = "no_data"
    else:
        price_match = "exact"

    agent_input: Dict[str, Any] = {
        **offer,
        "Slab": slab,
        "bl_apprvd": metrics["bl_apprvd"],
        "pur": metrics["pur"],
        "pur_retailer": metrics["pur_retailer"],
        "pur_wholesaler": metrics["pur_wholesaler"],
        "retail_ni": metrics["retail_ni"],
        "ni_retailer": metrics["ni_retailer"],
        "ni_wholesaler": metrics["ni_wholesaler"],
        "evidence_count": metrics.get("bucket_count", 0),
        "evidence_match": evidence_match,
        "q1": price["q1"],
        "median": price["median"],
        "q3": price["q3"],
        "price_match": price_match,
        "_override": _hard_override(offer, qty),
    }

    return {"agent_input": agent_input}


async def _classify(state: RetailState) -> RetailState:
    base_url = os.getenv("LLM_BASE_URL")
    api_key = os.getenv("LLM_API_KEY")
    model = os.getenv("LLM_MODEL")
    timeout = float(os.getenv("LLM_TIMEOUT", "60"))

    if not api_key or not model:
        raise RuntimeError("Missing LLM_API_KEY or LLM_MODEL")

    agent_input = state["agent_input"]
    override = agent_input.pop("_override", None)
    if override is not None:
        return {
            "system_prompt": "(skipped — hard override applied)",
            "user_message": "(skipped — hard override applied)",
            "raw_output": json.dumps(override, ensure_ascii=False),
            "result": override,
        }

    raw_prompt = _read_prompt()
    system_prompt = _render_template(raw_prompt, agent_input)

    _msg_keys = [
        "Display_id", "MCAT", "MCAT_id", "MCAT_Unit",
        "Qty", "Order_Value", "median", "Slab", "bl_apprvd", "pur",
        "pur_retailer", "pur_wholesaler", "retail_ni",
        "ni_retailer", "ni_wholesaler",
        "evidence_match", "evidence_count",
    ]
    user_text = "\n".join(f"{k}: {agent_input.get(k, '')}" for k in _msg_keys)

    last_exc: Exception | None = None
    response = None
    for attempt in range(1, _LLM_MAX_RETRIES + 2):
        try:
            try:
                llm = ChatOpenAI(
                    model=model,
                    api_key=api_key,
                    base_url=base_url,
                    timeout=timeout,
                    model_kwargs={
                        "response_format": {"type": "json_object"},
                        "reasoning_effort": "minimal",
                    },
                    extra_body={
                        "google": {"thinking_config": {"thinking_budget": 0}},
                    },
                )
                response = await llm.ainvoke([SystemMessage(content=system_prompt), HumanMessage(content=user_text)])
            except Exception:
                llm = ChatOpenAI(model=model, api_key=api_key, base_url=base_url, timeout=timeout)
                response = await llm.ainvoke([SystemMessage(content=system_prompt), HumanMessage(content=user_text)])
            break  # success
        except Exception as exc:
            last_exc = exc
            log.warning(
                "retail_agent LLM attempt %d/%d failed: %s: %s",
                attempt, _LLM_MAX_RETRIES + 1, type(exc).__name__, exc,
            )
            if attempt <= _LLM_MAX_RETRIES:
                await asyncio.sleep(2 ** (attempt - 1))
    else:
        raise RuntimeError(
            f"Retail LLM exhausted {_LLM_MAX_RETRIES + 1} attempts: {last_exc}"
        ) from last_exc

    raw_output = str(response.content)
    try:
        result = _clean_classifier_output(raw_output)
    except (ValueError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"Retail LLM output parse failed: {exc}. Raw: {raw_output[:500]!r}") from exc

    return {
        "system_prompt": system_prompt,
        "user_message": user_text,
        "raw_output": raw_output,
        "result": result,
    }


def _build_graph():
    graph = StateGraph(RetailState)
    graph.add_node("prepare_input", _prepare_input)
    graph.add_node("classify", _classify)
    graph.set_entry_point("prepare_input")
    graph.add_edge("prepare_input", "classify")
    graph.add_edge("classify", END)
    return graph.compile()


@lru_cache(maxsize=1)
def _compiled_graph():
    return _build_graph()


def _isq_quantity_missing_result(offer_id: str) -> Dict[str, Any]:
    return {
        "Display_id": offer_id,
        "Classification": "NON-RETAIL",
        "Classi_Score": 1.0,
        "Confidence": "High",
        "Override_Applied": "Yes — ISQ Quantity missing",
        "Reason": "Hard override: Quantity is missing from ISQ — cannot classify, treated as non-retail.",
    }


async def run_retail_agent(
    offer_id: str,
    buylead_response: Dict[str, Any],
    _trace: bool = False,
) -> Dict[str, Any]:
    from app.services.buylead_service import isq_has_quantity
    if not isq_has_quantity(buylead_response):
        result = _isq_quantity_missing_result(offer_id)
        if not _trace:
            return result
        return {
            "result": result,
            "agent_input": {"_override": "ISQ Quantity missing"},
            "raw_output": json.dumps(result, ensure_ascii=False),
            "system_prompt": "(skipped — ISQ Quantity missing)",
            "user_message": "(skipped — ISQ Quantity missing)",
        }

    from app.services.langfuse_service import get_langfuse_handler
    _h = get_langfuse_handler()
    _cfg = {"callbacks": [_h], "run_name": f"retail_agent:{offer_id}"} if _h else {}
    state = await _compiled_graph().ainvoke(
        {
            "offer_id": offer_id,
            "buylead_response": buylead_response,
        },
        config=_cfg,
    )
    if _trace:
        return {
            "result": state["result"],
            "agent_input": state.get("agent_input", {}),
            "raw_output": state.get("raw_output", ""),
            "system_prompt": state.get("system_prompt", ""),
            "user_message": state.get("user_message", ""),
        }
    return state["result"]
