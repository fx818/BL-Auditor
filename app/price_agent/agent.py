import json
import os
import re
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, TypedDict

from langchain_core.messages import HumanMessage, SystemMessage
from langchain_openai import ChatOpenAI
from langgraph.graph import END, StateGraph
from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
PROMPT_PATH = Path(__file__).resolve().with_name("prompt.md")
MCAT_DATA_PATH = ROOT_DIR / "mcat_data.xlsx"
EVIDENCE_DATA_PATH = ROOT_DIR / "evidence_data.xlsx"


class PriceState(TypedDict, total=False):
    offer_id: str
    buylead_response: Dict[str, Any]
    agent_input: Dict[str, Any]
    system_prompt: str
    user_message: str
    raw_output: str
    result: Dict[str, Any]


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    return "" if text in {"null", "undefined"} else text


def _to_number(value: Any, default: float = 0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _parse_jsonish(raw: Any) -> Any:
    if raw in (None, ""):
        return None
    if not isinstance(raw, str):
        raw = json.dumps(raw)
    raw = raw.replace('\\"', '"')
    parsed = json.loads(raw)
    if isinstance(parsed, str):
        parsed = json.loads(parsed)
    return parsed


_UNIT_SYNONYMS: Dict[str, str] = {
    # Tonne family — mcat uses TONNE, not TON
    "TON": "TONNE", "TONS": "TONNE", "TONE": "TONNE", "TONES": "TONNE",
    "TONNES": "TONNE", "TONNE": "TONNE",
    # Litre family
    "LITER": "LITRE", "LITERS": "LITRE", "LITRES": "LITRE", "LITRE": "LITRE",
    "LTR": "LITRE", "LTRS": "LITRE",
    # Piece family
    "PIECES": "PIECE", "PIECE": "PIECE", "PCS": "PIECE", "PC": "PIECE",
    # Box family — fixes "Boxes" → "BOX"
    "BOXES": "BOX", "BOX": "BOX",
    # Bottle family
    "BOTTLES": "BOTTLE", "BOTTLE": "BOTTLE",
    # Kilogram family
    "KILO": "KG", "KILOS": "KG", "KILOGRAM": "KG", "KILOGRAMS": "KG",
    "KGS": "KG", "KG": "KG",
    # Gram family — mcat uses GRAM, not G
    "GRAM": "GRAM", "GRAMS": "GRAM", "GMS": "GRAM", "GM": "GRAM", "G": "GRAM",
    # Bag, Jar, Set, Pack, Carton, Case, Pair, Roll
    "BAGS": "BAG", "BAG": "BAG",
    "JARS": "JAR", "JAR": "JAR",
    "SETS": "SET", "SET": "SET",
    "PACKS": "PACK", "PACK": "PACK",
    "CARTONS": "CARTON", "CARTON": "CARTON",
    "CASES": "CASE", "CASE": "CASE",
    "PAIRS": "PAIR", "PAIR": "PAIR",
    "ROLLS": "ROLL", "ROLL": "ROLL",
    # Meter family
    "METER": "METER", "METERS": "METER", "METRE": "METER", "METRES": "METER",
}


def _canonical_unit(value: Any) -> str:
    """Map a unit string (BL qty or evidence-file column) to a canonical
    uppercase form aligned with mcat's ``unit_display_name``.

    Strips leading numeric prefix (e.g. "50 Piece" → "Piece") and applies a
    synonym table for known typos/plurals. Multi-word phrases like "Carton
    Of 100 Pieces" pass through uppercased verbatim so they can match mcat
    entries such as "BOX OF 100 PIECES" if present.
    """
    raw = str(value or "").strip()
    raw = re.sub(r"^\d+\s*", "", raw).strip()
    if not raw:
        return ""
    upper = raw.upper()
    if " " in upper:
        return upper
    return _UNIT_SYNONYMS.get(upper, upper)


def _normalize_unit_from_qty(qty: str) -> str:
    return _canonical_unit(qty)


def _normalize_evidence_unit(raw_unit: Any) -> tuple[str, str]:
    canonical = _canonical_unit(raw_unit)
    if not canonical:
        return "no_unit", "No_Unit"
    return canonical, canonical


def _normalize_mcat_unit(value: Any) -> str:
    return "" if value is None else str(value).strip().upper()


def _extract_qty(qty_str: Any) -> float | None:
    if not qty_str:
        return None
    match = re.search(r"\d+", str(qty_str))
    return float(match.group(0)) if match else None


def _get_slab(qty: float | None) -> str:
    if qty is None or qty <= 0:
        return "no_slab"
    if qty <= 10:
        return "1-10"
    if qty <= 25:
        return "11-25"
    if qty <= 50:
        return "26-50"
    if qty <= 100:
        return "51-100"
    if qty <= 200:
        return "101-200"
    return "200+"


def _read_prompt() -> str:
    return PROMPT_PATH.read_text(encoding="utf-8")


def _render_template(template: str, data: Dict[str, Any]) -> str:
    def replace(match: re.Match[str]) -> str:
        key = match.group(1).strip()
        value = data.get(key, "")
        if isinstance(value, str):
            return value
        return json.dumps(value, ensure_ascii=False)

    return re.sub(r"{{\s*([^}]+)\s*}}", replace, template)


def _build_offer_source(offer_id: str, buylead_response: Dict[str, Any]) -> Dict[str, Any]:
    data = buylead_response.get("RESPONSE", {}).get("DATA", {})
    isq_info = data.get("ENRICHMENTINFO")

    buyer_profile = [
        {
            "eto_ofr_buyer_sell_mcats": data.get("ETO_OFR_BUYER_SELL_MCATS"),
            "eto_ofr_buyer_prime_mcats": data.get("ETO_OFR_BUYER_PRIME_MCATS"),
        }
    ]

    bl_card_data: List[Dict[str, Any]] = []
    for product in data.get("PRODUCTS_ENQUIRED") or []:
        if isinstance(product, dict):
            bl_card_data.append(product)

    return {
        "Display_id": data.get("ETO_OFR_DISPLAY_ID") or offer_id,
        "Title": data.get("ETO_OFR_TITLE") or "",
        "MCAT": data.get("PRIME_MCAT_NAME") or data.get("ETO_OFR_GLCAT_MCAT_NAME") or "",
        "MCAT_id": data.get("FK_GLCAT_MCAT_ID") or "",
        "ISQ_info": isq_info,
        "BL_Type": data.get("BY_LEAD_TYPE") or data.get("FK_ETO_OFR_TYPE_ID"),
        "Buyer_profile": buyer_profile,
        "BL_card_data": bl_card_data,
    }


def _bl_detail_and_keys(source: Dict[str, Any]) -> Dict[str, Any]:
    isq_parsed: Dict[str, str] = {}
    qty = ""
    order_value = ""

    try:
        parsed = _parse_jsonish(source.get("ISQ_info"))
        if isinstance(parsed, dict):
            parsed = parsed.get("1", [])
        if isinstance(parsed, list):
            for question in parsed:
                key = str(question.get("DESC") or "").strip() if isinstance(question, dict) else ""
                value = str(question.get("RESPONSE") or "").strip() if isinstance(question, dict) else ""
                if not key:
                    continue
                lowered = key.lower()
                if "quantity" in lowered:
                    qty = value
                    continue
                if "order value" in lowered:
                    order_value = value
                    continue
                isq_parsed[key] = value
    except Exception:
        pass

    unit = _normalize_unit_from_qty(qty)
    mcat_id = source.get("MCAT_id") or ""
    mcat_unit = f"{mcat_id}-{unit}" if mcat_id and unit else ""

    retail_flag = "No"
    try:
        if int(source.get("BL_Type") or 0) in {1, 3, 5, 6}:
            retail_flag = "Yes"
    except (TypeError, ValueError):
        pass

    buyer_obj: Dict[str, Any] = {}
    try:
        parsed_buyer = _parse_jsonish(source.get("Buyer_profile"))
        if isinstance(parsed_buyer, list):
            buyer_obj = parsed_buyer[0] or {}
        elif isinstance(parsed_buyer, dict):
            values = list(parsed_buyer.values())
            buyer_obj = values[0] if values and isinstance(values[0], dict) else parsed_buyer
    except Exception:
        pass

    bl_card = []
    try:
        parsed_card = _parse_jsonish(source.get("BL_card_data"))
        if isinstance(parsed_card, dict):
            parsed_card = [parsed_card]
        if isinstance(parsed_card, list):
            for card in parsed_card:
                if not isinstance(card, dict):
                    continue
                item_name = str(card.get("FK_PC_ITEM_NAME") or card.get("FK_PC_ITEM_DISPLAY_NAME") or "").strip()
                if not item_name:
                    continue
                price = str(card.get("PRODUCT_PRICE") or "").replace("\\u20b9", "₹").strip()
                bl_card.append({"Item Name": item_name, "Price": price})
    except Exception:
        bl_card = []

    return {
        "Display_id": source.get("Display_id"),
        "Title": source.get("Title"),
        "MCAT": source.get("MCAT"),
        "MCAT_id": mcat_id,
        "MCAT_Unit": mcat_unit,
        "Retail_Flag": retail_flag,
        "ISQ": isq_parsed,
        "Qty": qty,
        "Order_Value": order_value,
        "Buyer_profile": {
            "Sells": _clean(buyer_obj.get("eto_ofr_buyer_sell_mcats")) or "No Selling Activity",
            "Buys": _clean(buyer_obj.get("eto_ofr_buyer_prime_mcats")),
        },
        "BL_card": bl_card,
    }


_METRIC_FIELDS_FROM_EVIDENCE = (
    ("bl_apprvd", "bl_apprvd"),
    ("pur", "pur"),
    ("pur_retailer", "pur_retailer"),
    ("pur_wholesaler", "pur_wholesaler"),
    ("retail_ni", "retail_ni"),
    ("ni_retailer", "ret_ni_cnt_retailer"),
    ("ni_wholesaler", "ret_ni_cnt_wholesaler"),
)


def _empty_evidence_metrics() -> Dict[str, Any]:
    return {
        "bl_apprvd": 0, "pur": 0, "pur_retailer": 0, "pur_wholesaler": 0,
        "retail_ni": 0, "ni_retailer": 0, "ni_wholesaler": 0,
        "bucket_count": 0,
    }


@lru_cache(maxsize=1)
def _load_evidence_metrics() -> tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    """Build two evidence indexes from evidence_data.xlsx:
       - by_slab:  key = f"{norm_unit}_{slab}" → aggregated metrics for that exact bucket
       - by_unit:  key = f"{norm_unit}"        → metrics aggregated across all slabs for that (mcat,unit)
    Each metrics dict includes ``bucket_count`` = number of source rows aggregated.
    """
    wb = load_workbook(EVIDENCE_DATA_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))

    def _first(name: str) -> int:
        for i, h in enumerate(headers):
            if h == name:
                return i
        return -1

    mcat_id_idx = _first("glcat_mcat_id")
    mcat_name_idx = _first("glcat_mcat_name")
    qty_idx = _first("eto_ofr_qty")
    qty_unit_idx = _first("eto_ofr_qty_unit")
    field_idxs = {out: _first(src) for out, src in _METRIC_FIELDS_FROM_EVIDENCE}

    by_slab: Dict[str, Dict[str, Any]] = {}
    by_unit: Dict[str, Dict[str, Any]] = {}

    for row in rows:
        glcat_mcat_id = (row[mcat_id_idx] if mcat_id_idx >= 0 else None) or "Unknown_ID"
        glcat_mcat_name = (row[mcat_name_idx] if mcat_name_idx >= 0 else None) or "Unknown"
        raw_unit = (row[qty_unit_idx] if qty_unit_idx >= 0 else None) or ""
        qty = _to_number(row[qty_idx] if qty_idx >= 0 else 0, 0)

        if qty <= 0 or not str(raw_unit).strip():
            unit, display_unit, slab = "no_unit", "No_Unit", "no_slab"
        else:
            unit, display_unit = _normalize_evidence_unit(raw_unit)
            slab = _get_slab(qty)

        mcat_id_text = str(int(glcat_mcat_id)) if isinstance(glcat_mcat_id, float) else str(glcat_mcat_id)
        norm_unit = _normalize_mcat_unit(f"{mcat_id_text}-{display_unit}")
        slab_key = f"{norm_unit}_{slab}"
        unit_key = norm_unit

        if slab_key not in by_slab:
            entry = _empty_evidence_metrics()
            entry["glcat_mcat_id"] = mcat_id_text
            entry["glcat_mcat_name"] = glcat_mcat_name
            entry["eto_ofr_qty_unit"] = unit
            entry["slab"] = slab
            entry["MCAT_Unit"] = norm_unit
            by_slab[slab_key] = entry

        if unit_key not in by_unit:
            entry = _empty_evidence_metrics()
            entry["glcat_mcat_id"] = mcat_id_text
            entry["glcat_mcat_name"] = glcat_mcat_name
            entry["eto_ofr_qty_unit"] = unit
            entry["MCAT_Unit"] = norm_unit
            by_unit[unit_key] = entry

        for out_field, idx in field_idxs.items():
            if idx < 0:
                continue
            val = _to_number(row[idx])
            by_slab[slab_key][out_field] += val
            by_unit[unit_key][out_field] += val

        by_slab[slab_key]["bucket_count"] += 1
        by_unit[unit_key]["bucket_count"] += 1

    wb.close()
    return by_slab, by_unit


@lru_cache(maxsize=1)
def _load_price_index() -> Dict[str, Dict[str, float]]:
    """Build map from canonical `{mcat_id}-{UNIT}` → q1/median/q3 once.

    mcat_data.xlsx has duplicate 'median'/'q3' header names. This uses the
    FIRST occurrence (positions 8, 9, 10 in current schema) to stay
    deterministic if the duplicate columns ever diverge.
    """
    wb = load_workbook(MCAT_DATA_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))

    def _first(name: str) -> int:
        for i, h in enumerate(headers):
            if h == name:
                return i
        return -1

    mcat_idx = _first("fk_glcat_mcat_id")
    unit_idx = _first("unit_display_name")
    q1_idx = _first("q1")
    median_idx = _first("median")
    q3_idx = _first("q3")

    index: Dict[str, Dict[str, float]] = {}
    if -1 in (mcat_idx, unit_idx, q1_idx, median_idx, q3_idx):
        wb.close()
        return index

    for row in rows:
        mcat_id = row[mcat_idx]
        unit = row[unit_idx]
        if mcat_id in (None, "") or unit in (None, ""):
            continue
        mcat_id_text = str(int(mcat_id)) if isinstance(mcat_id, float) else str(mcat_id)
        key = _normalize_mcat_unit(f"{mcat_id_text}-{unit}")
        if key in index:
            continue
        index[key] = {
            "q1": _to_number(row[q1_idx]),
            "median": _to_number(row[median_idx]),
            "q3": _to_number(row[q3_idx]),
        }
    wb.close()
    return index


def _find_price_data(norm_unit: str) -> Dict[str, float] | None:
    if not norm_unit:
        return None
    return _load_price_index().get(norm_unit)


def _clean_price_output(raw: str) -> Dict[str, Any]:
    text = (raw or "").replace("```json", "").replace("```", "").strip()
    if not text:
        raise ValueError("LLM returned empty response")
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{[\s\S]*\}", text)
        if not match:
            raise ValueError(f"No JSON object found in LLM output: {text[:300]}")
        return json.loads(match.group(0))


async def _prepare_input(state: PriceState) -> PriceState:
    offer_id = state["offer_id"]
    source = _build_offer_source(offer_id, state["buylead_response"])
    offer = _bl_detail_and_keys(source)

    qty = _extract_qty(offer.get("Qty"))
    slab = _get_slab(qty)
    norm_unit = _normalize_mcat_unit(offer.get("MCAT_Unit"))

    by_slab, by_unit = _load_evidence_metrics()
    evidence_match = "no_data"
    evidence_count = 0
    slab_key = f"{norm_unit}_{slab}" if (norm_unit and slab and slab != "no_slab") else ""
    if slab_key and slab_key in by_slab:
        evidence_match = "exact"
        evidence_count = by_slab[slab_key].get("bucket_count", 0)
    elif norm_unit and norm_unit in by_unit:
        evidence_match = "unit_only"
        evidence_count = by_unit[norm_unit].get("bucket_count", 0)

    price = _find_price_data(norm_unit)
    if price is None:
        price = {"median": 0, "q3": 0}
        price_match = "no_data"
    else:
        price_match = "exact"

    agent_input: Dict[str, Any] = {
        "Display_id": offer.get("Display_id"),
        "Title": offer.get("Title"),
        "MCAT": offer.get("MCAT"),
        "Qty": offer.get("Qty"),
        "Order_Value": offer.get("Order_Value"),
        "BL_card": offer.get("BL_card"),
        "median": price.get("median", 0),
        "q3": price.get("q3", 0),
        "price_match": price_match,
        "evidence_match": evidence_match,
        "evidence_count": evidence_count,
    }

    return {"agent_input": agent_input}


async def _price_classify(state: PriceState) -> PriceState:
    base_url = os.getenv("PRICE_LLM_BASE_URL")
    api_key = os.getenv("PRICE_LLM_API_KEY")
    model = os.getenv("PRICE_LLM_MODEL")
    timeout = float(os.getenv("PRICE_LLM_TIMEOUT", "60"))

    if not api_key or not model:
        raise RuntimeError("Missing PRICE_LLM_API_KEY or PRICE_LLM_MODEL")

    agent_input = state["agent_input"]
    raw_prompt = _read_prompt()
    system_prompt = _render_template(raw_prompt, agent_input)

    user_text = "\n".join([
        f"Display_id: {agent_input.get('Display_id', '')}",
        f"MCAT: {agent_input.get('MCAT', '')}",
        f"Qty: {agent_input.get('Qty', '')}",
        f"Order_Value: {agent_input.get('Order_Value', '')}",
        f"BL_card: {json.dumps(agent_input.get('BL_card', []), ensure_ascii=False)}",
        f"median: {agent_input.get('median', '')}",
        f"q3: {agent_input.get('q3', '')}",
        f"price_match: {agent_input.get('price_match', '')}",
        f"evidence_match: {agent_input.get('evidence_match', '')}",
        f"evidence_count: {agent_input.get('evidence_count', '')}",
    ])

    try:
        llm = ChatOpenAI(
            model=model,
            api_key=api_key,
            base_url=base_url,
            timeout=timeout,
            model_kwargs={
                "response_format": {"type": "json_object"},
                "reasoning_effort": "minimal",
            },
            extra_body={
                "google": {"thinking_config": {"thinking_budget": 0}},
            },
        )
        response = await llm.ainvoke([SystemMessage(content=system_prompt), HumanMessage(content=user_text)])
    except Exception:
        llm = ChatOpenAI(model=model, api_key=api_key, base_url=base_url, timeout=timeout)
        response = await llm.ainvoke([SystemMessage(content=system_prompt), HumanMessage(content=user_text)])

    raw_output = str(response.content)
    try:
        result = _clean_price_output(raw_output)
    except (ValueError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"Price LLM output parse failed: {exc}. Raw: {raw_output[:500]!r}") from exc

    return {
        "system_prompt": system_prompt,
        "user_message": user_text,
        "raw_output": raw_output,
        "result": result,
    }


def _build_graph():
    graph = StateGraph(PriceState)
    graph.add_node("prepare_input", _prepare_input)
    graph.add_node("price_classify", _price_classify)
    graph.set_entry_point("prepare_input")
    graph.add_edge("prepare_input", "price_classify")
    graph.add_edge("price_classify", END)
    return graph.compile()


@lru_cache(maxsize=1)
def _compiled_graph():
    return _build_graph()


async def run_price_agent(
    offer_id: str,
    buylead_response: Dict[str, Any],
    _trace: bool = False,
) -> Dict[str, Any]:
    state = await _compiled_graph().ainvoke(
        {
            "offer_id": offer_id,
            "buylead_response": buylead_response,
        }
    )
    if _trace:
        return {
            "result": state["result"],
            "agent_input": state.get("agent_input", {}),
            "raw_output": state.get("raw_output", ""),
            "system_prompt": state.get("system_prompt", ""),
            "user_message": state.get("user_message", ""),
        }
    return state["result"]
